import contextlib
import copy
import datetime

from difflib import SequenceMatcher

import logging

import os
import re

import numpy
import openpyxl

if __name__ == "__main__":
    import dirsetup

from datum import frame

from cypy.vectorpy import strtype

class header():
    """It is a table of params, columns are fields."""

    def __init__(self,**kwargs):
        """parameters should be predefined, all entries must be string."""

        if len(kwargs)==0:
            raise ValueError("At least one field is required.")

        params = []
        fields = []
        fsizes = []

        for param,field in kwargs.items():

            params.append(param)

            if isinstance(field,list) or isinstance(field,tuple):
                field = [str(data) for data in field]
            else:
                field = [str(field)]

            fields.append(field)
            fsizes.append(len(field))

        if len(set(fsizes))!=1:
            raise ValueError("The lengths of field are not equal!")

        super().__setattr__("params",params)
        super().__setattr__("fields",fields)

    def extend(self,row):

        if len(self.params)!=len(row):
            raise ValueError("The lengths of 'fields' and 'row' are not equal!")

        if isinstance(row,list) or isinstance(row,tuple):
            toextend = header(**dict(zip(self.params,row)))
        elif isinstance(row,dict):
            toextend = header(**row)
        elif isinstance(row,header):
            toextend = row

        for param,field in toextend.items():
            self.fields[self.params.index(param)].extend(field)

    def __setattr__(self,key,vals):

        raise AttributeError(f"'Header' object has no attribute '{key}'.")

    def __getattr__(self,param):

        field = self.fields[self.params.index(param)]

        _field = [] # it may actually return numbers too.

        for value in field:

            try:
                _field.append(float(value))
            except ValueError:
                _field.append(value)

        if len(_field)==1:

            _field, = _field

        return _field

    def __getitem__(self,key):

        if not isinstance(key,str):
            raise TypeError("key must be string!")

        for row in self:
            if row[0].lower()==key.lower():
                break
        
        return header(**dict(zip(self.params,row)))

    def __repr__(self,comment=None):

        return self.__str__(comment)

    def __str__(self,comment=None):

        if len(self)==1:
            return repr(tuple(self.fields))

        if comment is None:
            comment = ""

        fstring = comment
        
        underline = []

        for param,field in zip(self.params,self.fields):

            field_ = list(field)
            field_.append(param)

            count_ = max([len(value) for value in field_])

            fstring += f"{{:<{count_}}}   "
            
            underline.append("-"*count_)

        fstring += "\n"

        text = fstring.format(*[parm.capitalize() for parm in self.params])
        
        text += fstring.format(*underline)
        
        for row in self:
            text += fstring.format(*row)

        return text

    def __iter__(self):

        return iter([row for row in zip(*self.fields)])

    def __len__(self):

        if isinstance(self.fields[0],str):
            return 1
        else:
            return len(self.fields[0])

    def items(self):

        return iter([(p,f) for p,f in zip(self.params,self.fields)])

class dirmaster():
    """Base directory class to manage files in the input & output directories."""

    def __init__(self,homedir=None,filedir=None,filepath=None):
        """Initializes base directory class with home & file directories."""

        self.set_homedir(homedir)

        if filepath is None and filedir is None:
            self.set_filedir(None)
        elif filepath is None and filedir is not None:
            self.set_filedir(filedir)
        elif filepath is not None and filedir is None:
            self.set_filedir(filepath)
            self.set_filepath(filepath)
        else:
            self.set_filedir(filedir)
            self.set_filepath(filepath)

    def set_homedir(self,path=None):
        """Sets the home directory to put outputs."""

        if path is None:
            path = os.getcwd()

        if not os.path.isabs(path):
            path = os.path.normpath(os.path.join(os.getcwd(),path))

        if not os.path.isdir(path):
            path = os.path.dirname(path)

        super().__setattr__("homedir",path)

    def set_filedir(self,path=None):
        """Sets the file directory to get inputs."""

        if path is None:
            path = self.homedir

        if not os.path.isabs(path):
            path = self.get_abspath(path,homeFlag=True)

        if not os.path.isdir(path):
            path = os.path.dirname(path)

        super().__setattr__("filedir",path)

    def set_filepath(self,path=None):
        """Sets the file path for the cases when it is working on a single file."""

        if path is None:
            return

        if os.path.isabs(path):
            self.set_filedir(path)
        else:
            path = self.get_abspath(path)

        if os.path.isdir(path):
            return

        super().__setattr__("filepath",path)

    def set_extension(self,path=None,extension=None):

        if path is None:
            path = self.filepath

        if extension is None:
            extension = ""

        basename = os.path.basename(path)

        rootname = os.path.splitext(basename)[0]

        basename = f"{rootname}{extension}"

        return os.path.normpath(os.path.join(self.filedir,basename))

    def get_abspath(self,path,homeFlag=False):
        """Returns absolute path for a given relative path."""

        if os.path.isabs(path):
            return path
        elif homeFlag:
            return os.path.normpath(os.path.join(self.homedir,path))
        else:
            return os.path.normpath(os.path.join(self.filedir,path))

    def file_exists(self,path,homeFlag=False):

        path = self.get_abspath(path,homeFlag=homeFlag)

        return os.path.exists(path)

    def get_dirpath(self,path,homeFlag=False):
        """Returns absolute directory path for a given relative path."""

        path = self.get_abspath(path,homeFlag=homeFlag)

        if os.path.isdir(path):
            return path
        else:
            return os.path.dirname(path)

    def get_fnames(self,path=None,homeFlag=False,prefix=None,extension=None,returnAbsFlag=False,returnDirsFlag=False):
        """Return directory(folder)/file names for a given relative path."""

        if path is None:
            path = self.filedir if homeFlag is False else self.homedir
        else:
            path = self.get_dirpath(path,homeFlag=homeFlag)

        fnames = os.listdir(path)

        fpaths = [self.get_abspath(fname,homeFlag=homeFlag) for fname in fnames]

        if returnDirsFlag:

            foldernames = [fname for (fname,fpath) in zip(fnames,fpaths) if os.path.isdir(fpath)]
            folderpaths = [fpath for fpath in fpaths if os.path.isdir(fpath)]

            if prefix is None:
                if returnAbsFlag:
                    return folderpaths
                else:
                    return foldernames
            else:
                if returnAbsFlag:
                    return [folderpath for (folderpath,foldername) in zip(folderpaths,foldernames) if foldername.startswith(prefix)]
                else:
                    return [foldername for foldername in foldernames if foldername.startswith(prefix)]
        else:

            filenames = [fname for (fname,fpath) in zip(fnames,fpaths) if not os.path.isdir(fpath)]
            filepaths = [fpath for fpath in fpaths if not os.path.isdir(fpath)]

            if prefix is None and extension is None:
                if returnAbsFlag:
                    return filepaths
                else:
                    return filenames
            elif prefix is None and extension is not None:
                if returnAbsFlag:
                    return [filepath for (filepath,filename) in zip(filepaths,filenames) if filename.endswith(extension)]
                else:
                    return [filename for filename in filenames if filename.endswith(extension)]
            elif prefix is not None and extension is None:
                if returnAbsFlag:
                    return [filepath for (filepath,filename) in zip(filepaths,filenames) if filename.startswith(prefix)]
                else:
                    return [filename for filename in filenames if filename.startswith(prefix)]
            else:
                if returnAbsFlag:
                    return [filepath for (filepath,filename) in zip(filepaths,filenames) if filename.startswith(prefix) and filename.endswith(extension)]
                else:
                    return [filename for filename in filenames if filename.startswith(prefix) and filename.endswith(extension)]

    @property
    def basename(self):

        return os.path.basename(self.filepath)

    @property
    def rootname(self):

        return os.path.splitext(self.basename)[0]

    @property
    def extension(self):

        return os.path.splitext(self.filepath)[1]

class dirview():

    def __init__(self,dirpath):

        self.dirpath = dirpath

    def draw(self,window,func=None):

        self.root = window

        self.scrollbar = tkinter.ttk.Scrollbar(self.root)

        self.tree = tkinter.ttk.Treeview(self.root,show="headings tree",selectmode="browse",yscrollcommand=self.scrollbar.set)

        self.tree.pack(side=tkinter.LEFT,expand=1,fill=tkinter.BOTH)

        self.scrollbar.pack(side=tkinter.LEFT,fill=tkinter.Y)

        self.scrollbar.config(command=self.tree.yview)

        self.tree.bind("<Button-1>",lambda event: self.set_path(func,event))

        self.refill()

    def refill(self):

        self.tree.heading("#0",text="")

        self.tree.delete(*self.tree.get_children())

        iterator = os.walk(self.dirpath)

        parents_name = []
        parents_link = []

        counter  = 0

        while True:

            try:
                root,dirs,frames = next(iterator)
            except StopIteration:
                break

            if counter==0:
                dirname = os.path.split(root)[1]
                self.tree.heading("#0",text=dirname,anchor=tkinter.W)
                parents_name.append(root)
                parents_link.append("")
            
            parent = parents_link[parents_name.index(root)]

            for directory in dirs:
                link = self.tree.insert(parent,'end',iid=counter,text=directory)
                counter += 1

                parents_name.append(os.path.join(root,directory))
                parents_link.append(link)

            for file in frames:            
                self.tree.insert(parent,'end',iid=counter,text=file)
                counter += 1

    def set_path(self,func=None,event=None):

        if event is not None:
            region = self.tree.identify("region",event.x,event.y)
        else:
            return

        if region!="tree":
            return

        item = self.tree.identify("row",event.x,event.y)

        path = self.tree.item(item)['text']

        while True:

            item = self.tree.parent(item)

            if item:
                path = os.path.join(self.tree.item(item)['text'],path)
            else:
                path = os.path.join(self.dirpath,path)
                break

        if func is not None:
            func(path)
        else:
            print(path)

class txtfile(dirmaster):

    def __init__(self,dataframe=None,**kwargs):

        super().__init__(**kwargs)

        if dataframe is None:
            self.frame = frame()
        else:
            self.frame = dataframe

        # self.header = header(
        #     heads=self.frame.heads,
        #     units=self.frame.units,
        #     infos=self.frame.infos,
        #     )

    def write(self,filepath,comment=None,**kwargs):
        """It writes text form of frame."""

        if comment is None:
            comment = "# "

        with open(filepath,"w",encoding='utf-8') as txtmaster:

            txtmaster.write(self.header.__str__(comment=comment))
            txtmaster.write(f"{comment}\n")
            txtmaster.write(self.frame.__str__(limit=self.frame.shape[0],comment=comment,**kwargs))

        # numpy.savetxt("data.txt",Z,fmt="%s",header=header,footer=footer,comments="")

    def writeb(self,filepath):
        """It writes binary form of frame."""

        for header,datacolumn in zip(self._headers,self._running):
            kwargs[header] = datacolumn

        numpy.savez_compressed(filepath,**kwargs)

        # numpy.savez_compressed('data.npz', a=A, b=B, c=C)

def loadtxt(filepath,**kwargs):
    """
    Returns an instance of textio.txtfile for the given filepath.
    
    Arguments:
        filepath {str} -- path to the given txt file

    Keyword Arguments:
        homedir {str} -- path to the home (output) directory
        filedir {str} -- path to the file (input) directory
    
    Returns:
        textio.txtfile -- an instance of textio.txtfile filled with LAS file text.

    """

    # It creates an empty textio.txtfile instance.
    nullfile = txtfile(filepath=filepath,**kwargs)

    # It reads LAS file and returns textio.lasfile instance.
    return txtworm(nullfile).item

class txtworm():

    def __init__(self,filepath,headline=None,comments="#",delimiter=None,skiprows=0):

        self.headline   = headline
        self.comments   = comments
        self.delimiter  = delimiter
        self.skiprows   = skiprows
        
        with loadtxt.textopen(filepath) as filemaster:
            dataframe = self.text(filemaster)

        self.frame = dataframe

    def seekrow(self,filemaster,row):

        filemaster.seek(0)

        countrows = 0

        while True:

            if countrows >= row:
                break

            next(filemaster)

            countrows += 1

    def heads(self,filemaster):

        if self.headline is None:
            return key2column(size=self.numcols,dtype='str').vals

        self.seekrow(filemaster,self.headline)

        line = next(filemaster).strip()

        if self.delimiter is None:
            heads_ = re.sub(r"\s+"," ",line).split(" ")
        else:
            heads_ = line.split(self.delimiter)

        return heads_

    def types(self,filemaster):

        self.seekrow(filemaster,self.skiprows)

        while True:

            line = next(filemaster).strip()

            if len(line)<1:
                continue
            
            if line.startswith(self.comments):
                continue

            break

        if self.delimiter is None:
            row = re.sub(r"\s+"," ",line).split(" ")
        else:
            row = line.split(self.delimiter)

        return strtype(row)

    def text(self,filemaster):

        types = self.types(filemaster)

        if self.headline is None:
            self.numcols = len(types)

        dtypes = [numpy.dtype(type_) for type_ in types]

        floatFlags = [True if type_ is float else False for type_ in types]

        self.seekrow(filemaster,self.skiprows)

        if all(floatFlags):
            cols = numpy.loadtxt(filemaster,comments=self.comments,delimiter=self.delimiter,unpack=True)
        else:
            cols = numpy.loadtxt(filemaster,comments=self.comments,delimiter=self.delimiter,unpack=True,dtype=str)

        heads = self.heads(filemaster)

        running = [column(col,head=head,dtype=dtype) for col,head,dtype in zip(cols,heads,dtypes)]

        return frame(*running)

    @contextlib.contextmanager
    def textopen(filepath):

        if hasattr(filepath,'read'):
            yield filepath
        else:
            self.set_filepath(filepath)
            filemaster = open(self.filepath,"r")
            try:
                yield filemaster
            finally:
                filemaster.close()

class txtbatch(dirmaster):

    def __init__(self):

        pass

class xlsheet(dirmaster):

    def __init__(self,*args,**kwargs):

        super().__init__(*args,**kwargs)

    def write(self,filepath,title):

        wb = openpyxl.Workbook()

        sheet = wb.active

        if title is not None:
            sheet.title = title

        for line in running:
            sheet.append(line)

        wb.save(filepath)

def loadxlsheet(filepath,sheetname,**kwargs):

    pass

class xlworm(xlsheet):

    def __init__(self,filepath,homedir=None,filedir=None,**kwargs):

        with xlbatch.xlopen(filepath) as book:

            for sheet in sheets:

                print("Loading {} {}".format(filepath,sheet))

                dataframe = self.load(book,sheet,**kwargs)

                self.frames.append(dataframe)

    def load(self,book,sheet,sheetsearch=False,min_row=1,min_col=1,max_row=None,max_col=None,hrows=0):
        """It reads provided excel worksheet and returns it as a frame."""

        dataframe = frame()

        if sheet is None:
            sheetname = book.sheetnames[0]
        elif isinstance(sheet,int):
            sheetname = book.sheetnames[sheet]
        elif isinstance(sheet,str) and sheetsearch:
            srcscores = [SequenceMatcher(None,page,sheet).ratio() for page in book.sheetnames]
            sheetname = book.sheetnames[srcscores.index(max(srcscores))]
        elif isinstance(sheet,str):
            if sheet in book.sheetnames:
                sheetname = sheet
            else:
                raise ValueError(f"'{sheet}' could not be found in the xlbook, try sheetsearch=True.")
        else:
            raise TypeError(f"Expected sheet is either none, int or str, but the input type is {type(sheet[1])}.")

        # frame.sheetname = sheetname

        rows = book[sheetname].iter_rows(
            min_row=min_row,max_row=max_row,
            min_col=min_col,max_col=max_col,
            values_only=True)

        cols = []
    
        for index,row in enumerate(rows):
            if index==0:
                [cols.append([]) for _ in row]
            if any(row):
                [col.append(cell) for cell,col in zip(row,cols)]

        heads_ = key2column(size=(len(cols)+min_col-1),dtype='str')
        heads_ = heads_.vals[min_col-1:]

        for index,col in enumerate(cols):
            if any(col):
                info = " ".join([str(cell) for cell in col[:hrows] if cell is not None])
                datacolumn = column(col[hrows:],head=heads_[index],info=info)
                dataframe._setup(datacolumn)

        return dataframe

    @contextlib.contextmanager
    def xlopen(filepath):
        xlbook = openpyxl.load_workbook(filepath,read_only=True,data_only=True)
        try:
            yield xlbook
        finally:
            xlbook._archive.close()

class xlbatch(dirmaster):

    def __init__(self,filepaths=None,sheetnames=None,**kwargs):

        super().__init__(**kwargs)

        self.frames = []

        self.loadall(filepaths,sheetnames,**kwargs)

    def load(self,filepaths,sheetnames=None,**kwargs):
        """It add frames from the excel worksheet provided with filepaths and sheetnames."""

        if filepaths is None:
            return
        elif type(filepaths) is str:
            filepaths = (filepaths,)
        else:
            filepaths = tuple(filepaths)

        if sheetnames is None:
            sheets = (sheetnames,)
        elif type(sheetnames) is str:
            sheets = (sheetnames,)
        else:
            sheets = tuple(sheetnames)

        for filepath in filepaths:

            filepath = self.get_abspath(filepath)

            dataframe = loadbook(filepath,**kwargs)

            self.frames.append(dataframe)

            logging.info(f"Loaded {filepath} as expected.")

    def merge(self,cols=None,idframes=None,infosearch=False):
        """It merges all the frames as a single frame under the Excel class."""

        framemerged = frame()

        if cols is None:
            pass
        elif type(cols) is int:
            cols = (cols,)
        elif type(cols) is str:
            cols = (cols,)
        else:
            cols = tuple(cols)

        if idframes is None:
            idframes = range(len(self.frames))
        elif type(idframes) is int:
            idframes = (idframes,)
        elif type(idframes) is str:
            idframes = (idframes,)
        else:
            idframes = tuple(idframes)

        for idframe in idframes:

            frame = self.frames[idframe]

            if cols is None:
                datacolumns = [datacolumn for datacolumn in dataframe.running]
            elif not infosearch:
                datacolumns = []
                for index in tuple(cols):
                    if type(index) is int:
                        datacolumns.append(dataframe.running[index])
                    elif type(index) is str:
                        datacolumns.append(dataframe[index])
                    else:
                        raise TypeError(f"cols type should be either int or str, not {type(index)}")
            else:
                datacolumns = []
                for index in tuple(cols):
                    if type(index) is str:
                        scores = [SequenceMatcher(None,info,index).ratio() for info in frame.infos]
                        datacolumns.append(dataframe.running[scores.index(max(scores))])
                    else:
                        raise TypeError(f"cols type should be str, not {type(index)}")

            framemerged._append(*datacolumns)

        return framemerged

if __name__ == "__main__":

    import unittest

    from tests.textio import header

    unittest.main(header)