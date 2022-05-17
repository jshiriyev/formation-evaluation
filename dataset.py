import calendar

from datetime import datetime
from dateutil import parser
from dateutil import relativedelta

from difflib import SequenceMatcher

import logging

import math
import os
import re

import numpy as np
import openpyxl as opxl
import lasio

if __name__ == "__main__":
    import setup

class DirBase():
    """Base directory class to manage files in the input & output directories."""

    def __init__(self,homedir=None,filedir=None):
        """Initializes base directory class with home & file directories."""

        self.set_homedir(homedir)
        self.set_filedir(filedir)

    def set_homedir(self,path=None):
        """Sets home directory to put outputs."""

        if path is None:
            path = os.getcwd()
        elif not os.path.isdir(path):
            path = os.path.dirname(path)

        if os.path.isabs(path):
            self.homedir = path
        else:
            self.homedir = os.path.normpath(os.path.join(os.getcwd(),path))

    def set_filedir(self,path=None):
        """Sets file directory to get inputs."""

        if path is None:
            path = self.homedir
        elif not os.path.isdir(path):
            path = os.path.dirname(path)

        if os.path.isabs(path):
            self.filedir = path
        else:
            self.filedir = os.path.normpath(os.path.join(self.homedir,path))

    def get_abspath(self,path,homeFlag=False):
        """Returns absolute path for a given relative path."""

        if os.path.isabs(path):
            return path
        elif homeFlag:
            return os.path.normpath(os.path.join(self.homedir,path))
        else:
            return os.path.normpath(os.path.join(self.filedir,path))

    def get_dirpath(self,path,homeFlag=False):
        """Returns absolute directory path for a given relative path."""

        path = self.get_abspath(path,homeFlag=homeFlag)

        if os.path.isdir(path):
            return path
        else:
            return os.path.dirname(path)

    def get_fnames(self,path=None,prefix=None,extension=None,returnAbsFlag=False,returnDirsFlag=False):
        """Return directory(folder)/file names for a given relative path."""

        if path is None:
            path = self.filedir
        else:
            path = self.get_dirpath(path)

        fnames = os.listdir(path)

        fpaths = [self.get_abspath(fname,homeFlag=False) for fname in fnames]

        if returnDirsFlag:

            foldernames = [fname for (fname,fpath) in zip(fnames,fpaths) if os.path.isdir(fpath)]
            folderpaths = [fpath for fpath in fpaths if os.path.isdir(fpath)]

            if prefix is None:
                if returnAbsFlag:
                    return folderpaths
                else:
                    return foldernames
            else:
                if returnAbsFlag:
                    return [folderpath for (folderpath,foldername) in zip(folderpaths,foldernames) if foldername.startswith(prefix)]
                else:
                    return [foldername for foldername in foldernames if foldername.startswith(prefix)]
        else:

            filenames = [fname for (fname,fpath) in zip(fnames,fpaths) if not os.path.isdir(fpath)]
            filepaths = [fpath for fpath in fpaths if not os.path.isdir(fpath)]

            if prefix is None and extension is None:
                if returnAbsFlag:
                    return filepaths
                else:
                    return filenames
            elif prefix is None and extension is not None:
                if returnAbsFlag:
                    return [filepath for (filepath,filename) in zip(filepaths,filenames) if filename.endswith(extension)]
                else:
                    return [filename for filename in filenames if filename.endswith(extension)]
            elif prefix is not None and extension is None:
                if returnAbsFlag:
                    return [filepath for (filepath,filename) in zip(filepaths,filenames) if filename.startswith(prefix)]
                else:
                    return [filename for filename in filenames if filename.startswith(prefix)]
            else:
                if returnAbsFlag:
                    return [filepath for (filepath,filename) in zip(filepaths,filenames) if filename.startswith(prefix) and filename.endswith(extension)]
                else:
                    return [filename for filename in filenames if filename.startswith(prefix) and filename.endswith(extension)]

class DataFrame(DirBase):
    """It stores equal-size one-dimensional numpy arrays in a list."""

    def __init__(self,*args,**kwargs):
        """Initializes DataFrame with headers & running and parent class DirBase."""

        super().__init__(**kwargs)

        if len(args)==0:
            DataFrame.set_headers(self,colnum=0,init=True)
        elif isinstance(args[0],np.ndarray):
            DataFrame.set_running(self,*args,init=True)
        else:
            DataFrame.set_headers(self,*args,init=True)

    def set_headers(self,*args,cols=None,colnum=None,init=False):
        """Set headers and running based on one or more inputs."""

        if isinstance(cols,int):
            cols = (cols,)

        if init:
            self._headers = []
            self._running = []

        colnum0 = len(self._headers)

        if cols is None and colnum is None:
            [self._headers.append(arg) for arg in args]
            [self._running.append(np.array([])) for arg in args]

        if colnum is not None:
            indices = range(colnum0,colnum0+colnum)
            [self._headers.append("Col #{}".format(index)) for index in indices]
            [self._running.append(np.array([])) for index in indices]

        if cols is not None:
            if len(args)!=len(cols):
                logging.critical("Length of cols is not equal to number of provided arguments.")
            for (index,arg) in zip(cols,args):
                self._headers[index] = arg      

        if len(self._running)!=len(self._headers):
            logging.warning("The DataFrame has headers and columns different in size.")

        rownum = np.array([column.size for column in self._running])

        if np.unique(rownum).size>1:
            logging.warning("The DataFrame has columns with different size.")

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def set_running(self,*args,cols=None,colnum=None,headers=None,init=False):
        """Set running and headers based on one or more inputs."""

        if isinstance(cols,int):
            cols = (cols,)

        if init:
            self._headers = []
            self._running = []

        colnum0 = len(self._running)

        if cols is None and colnum is None:

            if headers is None:
                indices = range(colnum0,colnum0+len(args))
                headers = ("Col #{}".format(index) for index in indices)
            elif not isinstance(headers,list) and not isinstance(headers,tuple):
                headers = (str(headers),)
            
            [self._headers.append(header) for header in headers]
            [self._running.append(arg) for arg in args]

        if colnum is not None:

            if headers is None:
                indices = range(colnum0,colnum0+colnum)
                headers = ("Col #{}".format(index) for index in indices)
            elif not isinstance(headers,list) and not isinstance(headers,tuple):
                headers = (str(headers),)
                
            [self._headers.append(header) for header in headers]
            [self._running.append(np.array([])) for index in indices]

        if cols is not None:

            if len(args)!=len(cols):
                logging.critical("Length of cols is not equal to number of provided arguments.")

            for (index,arg) in zip(cols,args):
                if isinstance(arg,np.ndarray):
                    self._running[index] = arg
                elif hasattr(arg,"__len__"):
                    self._running[index] = np.array(arg)
                else:
                    self._running[index] = np.array([arg])

            if headers is None:
                pass
            elif not isinstance(headers,list) and not isinstance(headers,tuple):
                headers = (str(headers),)
                self.set_headers(*headers,cols=cols)
            elif headers is not None:
                self.set_headers(*headers,cols=cols)

        if len(self._running)!=len(self._headers):
            logging.warning("The DataFrame has headers and columns different in size.")

        rownum = np.array([column.size for column in self._running])

        if np.unique(rownum).size!=1:
            logging.warning("The DataFrame has columns with different size.")

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def columns(self,cols=None,match=None,inplace=False,returnflag=True):
        """Set or returns columns in running"""

        if cols is None:
            colsnew = tuple(range(len(self._running)))
        elif isinstance(cols,int):
            colsnew = (cols,)
        elif isinstance(cols,str):
            colsnew = (self._headers.index(cols),)
        elif isinstance(cols,list) or isinstance(cols,tuple):
            if isinstance(cols[0],int):
                colsnew = cols
            elif isinstance(cols[0],str):
                colsnew = [self._headers.index(col) for col in cols]
            else:
                logging.critical(f"Expected cols is the list or tuple of integer or string; input, however, is {cols}")
        else:
            logging.critical(f"Other than None, expected cols is integer or string or their list or tuples; input, however, is {cols}")

        cols = colsnew

        if match is None:
            conditional = np.full(self._running[0].shape,True)
        else:
            conditional = self._running[match[0]]==match[1]

        if inplace:

            self._headers = [self._headers[index] for index in cols]
            self._running = [self._running[index][conditional] for index in cols]

            self.headers = self._headers
            self.running = [np.asarray(column) for column in self._running]

        else:

            if returnflag:
                if len(cols)==1:
                    return self._running[cols[0]]
                else:
                    return [self._running[index][conditional] for index in cols]
            else:
                self.headers = [self._headers[index] for index in cols]
                self.running = [np.asarray(self._running[index][conditional]) for index in cols]

    def additems(self,**kwargs):

        [setattr(self,key,value) for key,value in kwargs.items()]

    def addchilditems(self,parent,**kwargs):

        if not hasattr(self,parent):
            class Section(): pass
            setattr(self,parent,Section())

        [setattr(getattr(self,parent),key,value) for key,value in kwargs.items()]

    def texttocolumn(self,colID=None,header=None,deliminator=None,maxsplit=None):

        colID = colID if colID is not None else self._headers.index(header)

        header_string = self._headers[colID]
        # header_string = re.sub(deliminator+'+',deliminator,header_string)
        column_string = np.asarray(self._running[colID])

        headers = header_string.split(deliminator)
        columns = column_string[0].split(deliminator)

        if maxsplit is None:
            maxsplit = max(len(headers),len(columns))

        headers = header_string.split(deliminator,maxsplit=maxsplit-1)

        if maxsplit>len(headers):
            indices = range(maxsplit-len(headers))
            [headers.append("Col ##{}".format(index)) for index in indices]

        running = []

        for index,string in enumerate(column_string):

            # string = re.sub(deliminator+'+',deliminator,string)
            row = string.split(deliminator,maxsplit=maxsplit-1)

            if len(row)<maxsplit:
                indices = range(maxsplit-len(row))
                [row.append("") for index in indices]

            running.append(row)

        running = np.array(running,dtype=str).T

        self._headers.pop(colID)
        self._running.pop(colID)

        for header,column in zip(headers,running):
            self._headers.insert(colID,header)
            self._running.insert(colID,column)
            colID += 1

        # line = re.sub(r"[^\w]","",line)
        # line = "_"+line if line[0].isnumeric() else line
        # vmatch = np.vectorize(lambda x: bool(re.compile('[Ab]').match(x)))
        
        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def columntotext(self,colIDs=None,headers=None,header_new=None,fstring=None):

        if colIDs is None:
            colIDs = [self._headers.index(header) for header in headers]

        if fstring is None:
            fstring = ("{} "*len(colIDs)).strip()

        vprint = np.vectorize(lambda *args: fstring.format(*args))

        column_new = [np.asarray(self._running[index]) for index in colIDs]

        column_new = vprint(*column_new)

        if header_new is None:
            header_new = fstring.format(*[self._headers[index] for index in colIDs])

        self._headers.append(header_new)
        self._running.append(column_new)

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def astype(self,header_index=None,header=None,dtype=None,nonetozero=False,datestring=False,shiftmonths=0):

        def shifting(string):

            date = parser.parse(string)+relativedelta.relativedelta(months=shiftmonths)
            days = calendar.monthrange(date.year,date.month)[1]

            return datetime(date.year,date.month,days)

        def tryconvert(string):

            try:
                return dtype(str(string).replace(",","."))
            except ValueError:
                return np.nan

        if header_index is None:
            header_index = self._headers.index(header)

        if datestring:

            if dtype is None:
                if shiftmonths != 0:
                    vdate = np.vectorize(lambda x: shifting(x))
                else:
                    vdate = np.vectorize(lambda x: parser.parse(x))
            else:
                if shiftmonths != 0:
                    vdate = np.vectorize(lambda x: dtype(shifting(x)))
                else:
                    vdate = np.vectorize(lambda x: dtype(parser.parse(x)))
            
        else:

            if nonetozero:
                vdate = np.vectorize(lambda x: dtype(x) if x is not None else dtype(0))
            elif dtype==float:
                vdate = np.vectorize(lambda x: tryconvert(x))
            elif dtype==int:
                vdate = np.vectorize(lambda x: dtype(round(float(str(x).replace(",",".")))))
            else:
                vdate = np.vectorize(lambda x: dtype(x))
            
        self._running[header_index] = vdate(self._running[header_index])

        self.running[header_index] = np.asarray(self._running[header_index])

    def edit_nones(self,colIDs=None,headers=None,replace=None):

        if colIDs is None:
            if headers is None:
                colIDs = range(len(self._running))
            else:
                colIDs = [self._headers.index(header) for header in headers]

        for index in colIDs:

            column = self._running[index]

            row_indices = [row_index for row_index,val in enumerate(column) if val is None]

            for row_index in row_indices:
                if replace is None:
                    column[row_index] = column[row_index-1]
                else:
                    column[row_index] = replace

            self._running[index] = column
            self.running[index] = np.asarray(column)

    def edit_dates(self,colID=None,header=None):

        if colID is None:
            colID = self._headers.index(header)

        vdate = np.vectorize(lambda x: x if x is isinstance(x,datetime) else datetime.today())

        self._running[colID] = vdate(self._running[colID])

        self.running[colID] = np.asarray(self._running[colID])

    def edit_strings(self,colID=None,header=None,fstring=None,upper=False,lower=False,zfill=None):

        if colID is None:
            colID = self._headers.index(header)

        if fstring is None:
            fstring = "{}"

        if upper:
            case = lambda x: str(x).upper()
        elif lower:
            case = lambda x: str(x).lower()
        else:
            case = lambda x: str(x)

        if zfill is None:
            string = lambda x: case(x)
        else:
            string = lambda x: case(x).zfill(zfill)

        editor = np.vectorize(lambda x: fstring.format(string(x)))

        self._running[colID] = editor(self._running[colID])

        self.running[colID] = np.asarray(self._running[colID])

    def set_rows(self,rows,row_indices=None):
        
        for row in rows:

            if row_indices is None:
                for col_index,column in enumerate(self._running):
                    self._running[col_index] = np.append(column,row[col_index])
            else:
                for col_index, _ in enumerate(self._running):
                    self._running[col_index][row_indices] = row[col_index]

            self.running = [np.asarray(column) for column in self._running]

    def get_rows(self,row_indices=None,match=None):

        if row_indices is None:
            if match is None:
                row_indices = range(self._running[0].size)
            else:
                column_index,phrase = match
                conditional = self._running[column_index]==phrase
                row_indices = np.arange(self._running[0].size)[conditional]

        elif type(row_indices)==int:
            row_indices = [row_indices]

        rows = [[column[index] for column in self._running] for index in row_indices]
        
        return rows

    def del_rows(self,row_indices=None,noneColIndex=None,inplace=False):

        all_rows = np.array([np.arange(self._running[0].size)])

        if row_indices is None:
            row_indices = [index for index,val in enumerate(self.running[noneColIndex]) if val is None]
        
        row_indices = np.array(row_indices).reshape((-1,1))

        comp_mat = all_rows==row_indices

        keep_index = ~np.any(comp_mat,axis=0)

        if inplace:
            self._running = [column[keep_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[keep_index]) for column in self._running]
            
    def sort(self,header_indices=None,headers=None,reverse=False,inplace=False,returnFlag=False):

        if header_indices is None:
            header_indices = [self.headers.index(header) for header in headers]

        columns = [self._running[index] for index in header_indices]

        columns.reverse()

        sort_index = np.lexsort(columns)

        if reverse:
            sort_index = np.flip(sort_index)

        if inplace:
            self._running = [column[sort_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[sort_index]) for column in self._running]

        if returnFlag:
            return sort_index

    def filter(self,header_index=None,header=None,keywords=None,regex=None,year=None,inplace=False):

        ## It should allow multiple header indices filtering

        if header_index is None:
            header_index = self._headers.index(header)

        if keywords is not None:
            match_array = np.array(keywords).reshape((-1,1))
            match_index = np.any(self._running[header_index]==match_array,axis=0)
        elif regex is not None:
            match_vectr = np.vectorize(lambda x: bool(re.compile(regex).match(x)))
            match_index = match_vectr(self._running[header_index])
        elif year is not None:
            match_vectr = np.vectorize(lambda x: x.year==year)
            match_index = match_vectr(self._running[header_index].tolist())

        if inplace:
            self._running = [column[match_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[match_index]) for column in self._running]

    def unique(self,header_indices,inplace=False):

        Z = [self._running[index].astype(str) for index in header_indices]

        Z = np.array(Z,dtype=str).T

        _,row_indices = np.unique(Z,axis=0,return_index=True)

        if inplace:
            self._running = [column[row_indices] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[row_indices]) for column in self._running]

    def invert(self):

        self.running = [np.asarray(column) for column in self._running]

    def print(self,cols=None,rows=None,rlim=20):
        """It prints to the console limited number of rows with headers."""

        rlim = int(rlim)

        if cols is None:
            cols = list(range(len(self.running)))

        fstring = "{}\t"*len(cols)

        headers = [self.headers[col] for col in cols]
        unlines = ["-"*len(self.headers[col]) for col in cols]

        if rows is None:
            if self.running[0].size<=rlim:
                rows = list(range(self.running[0].size))
                print("\n",fstring.format(*headers))
                print(fstring.format(*unlines))
                for row in rows:
                    print(fstring.format(*[self.running[col][row] for col in cols]))
            else:
                rows1 = list(range(int(rlim/2)))
                rows2 = list(range(-1,-int(rlim/2)-1,-1))
                print("\nHEAD DATA...")
                print(fstring.format(*headers))
                print(fstring.format(*unlines))
                for row in rows1:
                    print(fstring.format(*[self.running[col][row] for col in cols]))
                print("\nFOOT DATA...")
                print(fstring.format(*headers))
                print(fstring.format(*unlines))
                for row in rows2:
                    print(fstring.format(*[self.running[col][row] for col in cols]))
        else:
            print("\n",fstring.format(*headers))
            print(fstring.format(*unlines))
            for row in rows:
                print(fstring.format(*[self.running[col][row] for col in cols]))

    def write(self,filepath,fstring=None,**kwargs):

        header_fstring = ("{}\t"*len(self._headers))[:-1]+"\n"

        if fstring is None:
            running_fstring = ("{}\t"*len(self._headers))[:-1]+"\n"
        else:
            running_fstring = fstring

        vprint = np.vectorize(lambda *args: running_fstring.format(*args))

        with open(filepath,"w",encoding='utf-8') as wfile:
            wfile.write(header_fstring.format(*self._headers))
            for line in vprint(*self._running):
                wfile.write(line)

# Collective Data Input/Output Classes

class RegText(DataFrame):

    def __init__(self,filepaths=None,**kwargs):

        super().__init__(**kwargs)

        self.frames = []

        self.add_frames(filepaths,**kwargs)

    def add_frames(self,filepaths,**kwargs):

        if filepaths is None:
            return
        if not isinstance(filepaths,list) and not isinstance(filepaths,tuple):
            filepaths = (filepaths,)

        for filepath in filepaths:
            self.frames.append(self.read(filepath,**kwargs))
            logging.info(f"Loaded {filepath} as expected.")

    def read(self,filepath,delimiter="\t",comments="#",skiprows=None,nondigitflag=False):

        filepath = self.get_abspath(filepath)

        frame = DataFrame(filedir=filepath)

        with open(filepath,mode="r",encoding="latin1") as text:

            if skiprows is None:
                skiprows = 0
                line = next(text).split("\n")[0]
                while not line.split(delimiter)[0].isdigit():
                    skiprows += 1
                    print(line)
                    line = next(text).split("\n")[0]
            else:
                for _ in range(skiprows):
                    line = next(text)

            line = next(text).split("\n")[0]

        if nondigitflag:
            row = line.split(delimiter)
            dtypes = [float if column.isdigit() else str for column in row]
            columns = []
            for index,dtype in enumerate(dtypes):
                column = np.loadtxt(filepath,dtype=dtype,delimiter=delimiter,skiprows=skiprows,usecols=[index],encoding="latin1")
                columns.append(column)
        else:
            data = np.loadtxt(filepath,comments="#",delimiter=delimiter,skiprows=skiprows,encoding="latin1")
            columns = [column for column in data.transpose()]

        frame.set_running(*columns,init=True)

        return frame

class LogASCII(DataFrame):

    def __init__(self,filepaths=None,**kwargs):

        super().__init__(**kwargs)

        self.frames = []

        self.add_frames(filepaths,**kwargs)

    def add_frames(filepaths,**kwargs):

        if filepaths is None:
            return
        if not isinstance(filepaths,list) and not isinstance(filepaths,tuple):
            filepaths = (filepaths,)

        for filepath in filepaths:
            self.frames.append(self.read(filepath,**kwargs))
            logging.info(f"Loaded {filepath} as expected.")

    def read(self,filepath,headers=None):

        filepath = self.get_abspath(filepath)

        datasection = "{}ASCII".format("~")

        skiprows = 1

        frame = DataFrame(filedir=filepath)

        with open(filepath,"r",encoding="latin1") as text:

            line = next(text).strip()

            while not line.startswith(datasection[:2]):

                if line.startswith("~"):

                    title = line[1:].split()[0].lower()

                    mnemonics,units,values,descriptions = [],[],[],[]

                elif len(line)<1:
                    pass

                elif line.startswith("#"):
                    pass

                elif line.find(".")<0:
                    pass

                elif line.find(":")<0:
                    pass

                elif line.index(".")>line.index(":"):
                    pass

                else:

                    line = re.sub(r'[^\x00-\x7F]+','',line)

                    mnemonic,rest = line.split(".",maxsplit=1)
                    rest,descrptn = rest.split(":",maxsplit=1)

                    if rest.startswith(" "):
                        unit = ""
                        value = rest.strip()
                    elif rest.endswith(" "):
                        value = ""
                        unit = rest.strip()
                    else:
                        unit,value = rest.split(" ",maxsplit=1)

                    mnemonics.append(mnemonic.strip())
                    units.append(unit.strip())

                    try:
                        value = int(value.strip())
                    except ValueError:
                        try:
                            value = float(value.strip())
                        except ValueError:
                            value = value.strip()

                    values.append(value)
                    descriptions.append(descrptn.strip())

                skiprows += 1

                line = next(text).strip()

                if line.startswith("~"):

                    if title=="version":

                        vnumb = "LAS {}".format(values[mnemonics.index("VERS")])
                        vinfo = descriptions[mnemonics.index("VERS")]
                        mtype = "Un-wrapped" if values[mnemonics.index("WRAP")]=="NO" else "Wrapped"
                        minfo = descriptions[mnemonics.index("WRAP")]

                        frame.additems(info=vnumb)
                        frame.additems(infodetail=vinfo)
                        frame.additems(mode=mtype)
                        frame.additems(modedetail=minfo)

                    elif title=="curve":

                        frame.set_headers(*mnemonics)
                        frame.additems(units=units)
                        frame.additems(details=descriptions)

                    else:

                        if len(mnemonics)>0:
                            frame.addchilditems(title,mnemonics=mnemonics)
                            frame.addchilditems(title,units=units)
                            frame.addchilditems(title,values=values)
                            frame.addchilditems(title,descriptions=descriptions)

        if headers is None:
            usecols = None
        else:
            usecols = [frame.headers.index(header) for header in headers]

        logdata = np.loadtxt(filepath,comments="#",skiprows=skiprows,usecols=usecols,encoding="latin1")

        nullvalue = frame.well.values[frame.well.mnemonics.index("NULL")]

        logdata[logdata==nullvalue] = np.nan

        columns = [column for column in logdata.transpose()]

        frame.set_running(*columns,cols=range(len(columns)))

        return frame

    def printwells(self,idframes=None):

        if idframes is None:
            idframes = tuple(range(len(self.frames)))
        elif isinstance(idframes,int):
            idframes = (idframes,)

        for index in idframes:

            frame = self.frames[index]

            print("\n\tWELL #{}".format(frame.well.values[frame.well.mnemonics.index("WELL")]))

            iterator = zip(frame.well.mnemonics,frame.well.units,frame.well.values,frame.well.descriptions)

            for mnem,unit,value,descr in iterator:
                print(f"{descr} ({mnem}):\t\t{value} [{unit}]")

    def printcurves(self,idframes=None,mnemonic_space=33,tab_space=8):

        if idframes is None:
            idframes = tuple(range(len(self.frames)))
        elif isinstance(idframes,int):
            idframes = (idframes,)

        for index in idframes:

            frame = self.frames[index]

            iterator = zip(frame.headers,frame.units,frame.details,frame.running)

            # file.write("\n\tLOG NUMBER {}\n".format(idframes))
            print("\n\tLOG NUMBER {}".format(index))

            for header,unit,detail,column in iterator:

                minXval = np.nanmin(column)
                maxXval = np.nanmax(column)

                tab_num = math.ceil((mnemonic_space-len(header))/tab_space)
                tab_spc = "\t"*tab_num if tab_num>0 else "\t"

                # file.write("Curve: {}{}Units: {}\tMin: {}\tMax: {}\tDescription: {}\n".format(
                #     curve.mnemonic,tab_spc,curve.unit,minXval,maxXval,curve.descr))
                print("Curve: {}{}Units: {}\tMin: {}\tMax: {}\tDescription: {}".format(
                    header,tab_spc,unit,minXval,maxXval,detail))

    def flip(self,idframes=None):

        if idframes is None:
            idframes = tuple(range(len(self.frames)))
        elif isinstance(idframes,int):
            idframes = (idframes,)

        for index in idframes:

            columns = [np.flip(column) for column in self.frames[index].running]

            self.frames[index].set_running(*columns,cols=range(len(columns)))
            
    def set_interval(self,top,bottom,idframes=None):

        if idframes is None:
            idframes = tuple(range(len(self.frames)))
        elif isinstance(idframes,int):
            idframes = (idframes,)

        self.top = top
        self.bottom = bottom

        self.gross_thickness = self.bottom-self.top

        for index in idframes:

            frame = self.frames[index]

            try:
                depth = frame.columns("MD")
            except ValueError:
                depth = frame.columns("DEPT")

            depth_cond = np.logical_and(depth>self.top,depth<self.bottom)

            columns = [column[depth_cond] for column in self.frame[index].running]

            self.frames[index].set_running(*columns,cols=range(len(columns)))

    def get_interval(self,top,bottom,fileID=None,curveID=None):

        returningList = []

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        for indexI in fileIDs:

            las = self.files[indexI]

            try:
                depth = las["MD"]
            except KeyError:
                depth = las["DEPT"]

            depth_cond = np.logical_and(depth>top,depth<bottom)

            if curveID is None:
                returningList.append(depth_cond)
            else:
                returningList.append(las.curves[curveID].data[depth_cond])

        return returningList

    def get_resampled(self,depthsR,depthsO,dataO):

        lowerend = depthsR<depthsO.min()
        upperend = depthsR>depthsO.max()

        interior = np.logical_and(~lowerend,~upperend)

        depths_interior = depthsR[interior]

        indices_lower = np.empty(depths_interior.shape,dtype=int)
        indices_upper = np.empty(depths_interior.shape,dtype=int)

        for index,depth in enumerate(depths_interior):

            diff = depthsO-depth

            indices_lower[index] = np.where(diff<0,diff,-np.inf).argmax()
            indices_upper[index] = np.where(diff>0,diff,np.inf).argmin()

        grads = (depths_interior-depthsO[indices_lower])/(depthsO[indices_upper]-depthsO[indices_lower])

        dataR = np.empty(depthsR.shape,dtype=float)

        dataR[lowerend] = np.nan
        dataR[interior] = dataO[indices_lower]+grads*(dataO[indices_upper]-dataO[indices_lower])
        dataR[upperend] = np.nan

        return dataR

    def resample(self,depthsFID=None,depthsR=None,fileID=None,curveID=None):

        """

        depthsFID:  The index of file id from which to take new depthsR
                    where new curve data will be calculated;

        depthsR:    The numpy array of new depthsR
                    where new curve data will be calculated;
        
        fileID:     The index of file to resample;
                    If None, all files will be resampled;
        
        curveID:    The index of curve in the las file to resample;
                    If None, all curves in the file will be resampled;
                    Else if fileID is not None, resampled data will be returned;

        """

        if depthsFID is not None:
            try:
                depthsR = self.files[depthsFID]["MD"]
            except KeyError:
                depthsR = self.files[depthsFID]["DEPT"]

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        for indexI in fileIDs:

            if depthsFID is not None:
                if indexI==depthsFID:
                    continue

            las = self.files[indexI]

            try:
                depthsO = las["MD"]
            except KeyError:
                depthsO = las["DEPT"]

            lowerend = depthsR<depthsO.min()
            upperend = depthsR>depthsO.max()

            interior = np.logical_and(~lowerend,~upperend)

            depths_interior = depthsR[interior]

            diff = depthsO-depths_interior.reshape((-1,1))

            indices_lower = np.where(diff<0,diff,-np.inf).argmax(axis=1)
            indices_upper = np.where(diff>0,diff,np.inf).argmin(axis=1)

            grads = (depths_interior-depthsO[indices_lower])/(depthsO[indices_upper]-depthsO[indices_lower])

            if curveID is None:
                las.curves[0].data = depthsR

            if curveID is None:
                curveIDs = range(1,len(las.curves))
            else:
                curveIDs = range(curveID,curveID+1)

            for indexJ in curveIDs:

                curve = las.curves[indexJ]

                dataR = np.empty(depthsR.shape,dtype=float)

                dataR[lowerend] = np.nan
                dataR[interior] = curve.data[indices_lower]+grads*(curve.data[indices_upper]-curve.data[indices_lower])
                dataR[upperend] = np.nan

                if curveID is None:
                    self.files[indexI].curves[indexJ].data = dataR
                elif fileID is not None:
                    return dataR

    def merge(self,fileIDs,curveNames):

        if isinstance(fileIDs,int):

            try:
                depth = self.files[fileIDs]["MD"]
            except KeyError:
                depth = self.files[fileIDs]["DEPT"]

            xvals1 = self.files[fileIDs][curveNames[0]]

            for curveName in curveNames[1:]:

                xvals2 = self.files[fileIDs][curveName]

                xvals1[np.isnan(xvals1)] = xvals2[np.isnan(xvals1)]

            return depth,xvals1

        elif np.unique(np.array(fileIDs)).size==len(fileIDs):

            if isinstance(curveNames,str):
                curveNames = (curveNames,)*len(fileIDs)

            depth = np.array([])
            xvals = np.array([])

            for (fileID,curveName) in zip(fileIDs,curveNames):

                try:
                    depth_loc = self.files[fileID]["MD"]
                except KeyError:
                    depth_loc = self.files[fileID]["DEPT"]

                xvals_loc = self.files[fileID][curveName]

                depth_loc = depth_loc[~np.isnan(xvals_loc)]
                xvals_loc = xvals_loc[~np.isnan(xvals_loc)]

                depth_max = 0 if depth.size==0 else depth.max()

                depth = np.append(depth,depth_loc[depth_loc>depth_max])
                xvals = np.append(xvals,xvals_loc[depth_loc>depth_max])

            return depth,xvals

    def write(self,filepath,mnemonics,data,fileID=None,units=None,descriptions=None,values=None):

        """
        filepath:       It will write a lasio.LASFile to the given filepath
        fileID:         The file index which to write to the given filepath
                        If fileID is None, new lasio.LASFile will be created

        kwargs:         These are mnemonics, data, units, descriptions, values
        """

        if fileID is not None:

            lasfile = self.files[fileID]

        else:

            lasfile = lasio.LASFile()

            lasfile.well.DATE = datetime.today().strftime('%Y-%m-%d %H:%M:%S')

            depthExistFlag = False

            for mnemonic in mnemonics:
                if mnemonic=="MD" or mnemonic=="DEPT":
                    depthExistFlag = True
                    break

            if not depthExistFlag:
                curve = lasio.CurveItem(
                    mnemonic="DEPT",
                    unit="",
                    value="",
                    descr="Depth index",
                    data=np.arange(data[0].size))
                lasfile.append_curve_item(curve)            

        for index,(mnemonic,datum) in enumerate(zip(mnemonics,data)):

            if units is not None:
                unit = units[index]
            else:
                unit = ""

            if descriptions is not None:
                description = descriptions[index]
            else:
                description = ""

            if values is not None:
                value = values[index]
            else:
                value = ""

            curve = lasio.CurveItem(mnemonic=mnemonic,data=datum,unit=unit,descr=description,value=value)

            lasfile.append_curve_item(curve)

        with open(filepath, mode='w') as filePathToWrite:
            lasfile.write(filePathToWrite)

class Excel(DataFrame):

    def __init__(self,homedir=None,filepath=None,headers=None):

        super().__init__(homedir=homedir,filepath=filepath,headers=headers)

        self.files = []
        self.books = []

    def add_frames(self,filepaths):

        [self.add_file(filepath) for filepath in filepaths]

    def add_file(self,filepath):

        filepath = self.get_filepathabs(filepath)

        self.files.append(opxl.load_workbook(filepath,read_only=True,data_only=True))

    def get_sheetname(self,keyword,fileID=0):

        # This method supposed to return sheetname similar to the keyword

        for sheetname in self.files[fileID].sheetnames:
            if sheetname[:len(keyword)]==keyword:
                return sheetname

        print("No good match for sheetname was found!")

    def set_headers(self,sheetname,row_index,min_col=None,max_col=None,fileID=0):

        headers = self.files[fileID][sheetname].iter_rows(
            min_row=row_index,min_col=min_col,
            max_row=row_index,max_col=max_col,
            values_only=True)

        headers = list(list(headers)[0])

        for index,header in enumerate(headers):
            headers[index] = re.sub(r"[^\w]","",header)

        super().set_headers(headers=headers,init=True)

    def read(self,sheetname=None,min_row=1,min_col=1,max_row=None,max_col=None,fileID=None):

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        if len(self._headers)==0:

            if len(self._running)==0: 
                super().set_headers(num_cols=max_col-min_col+1,init=True)
            else:
                super().set_headers(num_cols=max_col-min_col+1,init=False)

        for fileID in fileIDs:

            wsname = self.files[fileID].sheetnames[0] if sheetname is None else sheetname

            rows = self.files[fileID][wsname].iter_rows(
                min_row=min_row,min_col=min_col,
                max_row=max_row,max_col=max_col,
                values_only=True)

            rows = [row for row in list(rows) if any(row)]

            self.books.append(rows)

    def merge(self,header_rows=1):

        for i,book in enumerate(self.books):

            for j in range(1,header_rows):

                self.books[i][0] = [h0 if hj is None else hj for (h0,hj) in zip(book[0],book[j])]

            for j in range(1,header_rows):

                del self.books[i][1]

            col_indices = []

            for header in self._headers:

                scores = []

                for header_read in self.books[i][0]:

                    score = SequenceMatcher(None,header,header_read).ratio() if isinstance(header_read,str) else 0
                    
                    scores.append(score)

                col_indices.append(scores.index(max(scores)))

            for j,row in enumerate(self.books[i]):

                self.books[i][j] = [row[k] for k in col_indices]

            rows = [row for row in self.books[i][1:] if any(row)]

            self.set_rows(rows)

    def write(self,filepath,title):

        wb = opxl.Workbook()

        sheet = wb.active

        if title is not None:
            sheet.title = title

        for line in running:
            sheet.append(line)

        wb.save(filepath)

    def close(self,fileID=None):

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        for fileID in fileIDs:
            self.files[fileID]._archive.close()

class IrrText(DataFrame):

    def __init__(self,filepath,**kwargs):

        super().__init__(**kwargs)

        if filepath is not None:
            self.filepath = self.get_abspath(filepath,homeFlag=False)

    def set_batch(self,filepaths):

        self.filepaths = filepaths

    def read(self,skiprows=0,headerline=None,comment="--",endline="/",endfile="END"):

        # While looping inside the file it does not read lines:
        # - starting with comment phrase, e.g., comment = "--"
        # - after the end of line phrase, e.g., endline = "/"
        # - after the end of file keyword e.g., endfile = "END"

        if headerline is None:
            headerline = skiprows-1
        elif headerline<skiprows:
            headerline = headerline
        else:
            headerline = skiprows-1

        _running = []

        with open(self.filepath,"r") as text:

            for line in text:

                line = line.split('\n')[0].strip()

                line = line.strip(endline)

                line = line.strip()
                line = line.strip("\t")
                line = line.strip()

                if line=="":
                    continue

                if comment is not None:
                    if line[:len(comment)] == comment:
                        continue

                if endfile is not None:
                    if line[:len(endfile)] == endfile:
                        break

                _running.append([line])

        self.title = []

        for _ in range(skiprows):
            self.title.append(_running.pop(0))

        num_cols = len(_running[0])

        if skiprows==0:
            self.set_headers(num_cols=num_cols,init=False)
        elif skiprows!=0:
            self.set_headers(headers=self.title[headerline],init=False)

        nparray = np.array(_running).T

        self._running = [np.asarray(column) for column in nparray]

        self.running = [np.asarray(column) for column in self._running]

    def write(self):

        pass

class WSchedule(DataFrame):

    # KEYWORDS: DATES,COMPDATMD,COMPORD,WCONHIST,WCONINJH,WEFAC,WELOPEN 

    headers    = ["DATE","KEYWORD","DETAILS",]

    dates      = " {} / "#.format(date)
    welspecs   = " '{}'\t1*\t2* / "
    compdatop  = " '{}'\t1*\t{}\t{}\tMD\t{}\t2*\t0.14 / "#.format(wellname,top,bottom,optype)
    compdatsh  = " '{}'\t1*\t{}\t{}\tMD\t{} / "#.format(wellname,top,bottom,optype)
    compord    = " '{}'\tINPUT\t/ "#.format(wellname)
    prodhist   = " '{}'\tOPEN\tORAT\t{}\t{}\t{} / "#.format(wellname,oilrate,waterrate,gasrate)
    injhist    = " '{}'\tWATER\tOPEN\t{}\t7*\tRATE / "#.format(wellname,waterrate)
    wefac      = " '{}'\t{} / "#.format(wellname,efficiency)
    welopen    = " '{}'\tSHUT\t3* / "#.format(wellname)

    def __init__(self):

        pass

    def read(self):

        pass

    def set_subheaders(self,header_index=None,header=None,regex=None,regex_builtin="INC_HEADERS",title="SUB-HEADERS"):

        nparray = np.array(self._running[header_index])

        if regex is None and regex_builtin=="INC_HEADERS":
            regex = r'^[A-Z]+$'                         #for strings with only capital letters no digits
        elif regex is None and regex_builtin=="INC_DATES":
            regex = r'^\d{1,2} [A-Za-z]{3} \d{2}\d{2}?$'   #for strings with [1 or 2 digits][space][3 capital letters][space][2 or 4 digits], e.g. DATES

        vmatch = np.vectorize(lambda x: bool(re.compile(regex).match(x)))

        match_index = vmatch(nparray)

        firstocc = np.argmax(match_index)

        lower = np.where(match_index)[0]
        upper = np.append(lower[1:],nparray.size)

        repeat_count = upper-lower-1

        match_content = nparray[match_index]

        nparray[firstocc:][~match_index[firstocc:]] = np.repeat(match_content,repeat_count)

        self._headers.insert(header_index,title)
        self._running.insert(header_index,np.asarray(nparray))

        for index,column in enumerate(self._running):
            self._running[index] = np.array(self._running[index][firstocc:][~match_index[firstocc:]])

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def get_wells(self,wellname=None):

        pass

    def write(self):

        path = os.path.join(self.workdir,self.schedule_filename)

        with open(path,"w",encoding='utf-8') as wfile:

            welspec = schedule.running[1]=="WELSPECS"
            compdat = schedule.running[1]=="COMPDATMD"
            compord = schedule.running[1]=="COMPORD"
            prodhst = schedule.running[1]=="WCONHIST"
            injdhst = schedule.running[1]=="WCONINJH"
            wefffac = schedule.running[1]=="WEFAC"
            welopen = schedule.running[1]=="WELOPEN"

            for date in np.unique(schedule.running[0]):

                currentdate = schedule.running[0]==date

                currentcont = schedule.running[1][currentdate]

                wfile.write("\n\n")
                wfile.write("DATES\n")
                wfile.write(self.schedule_dates.format(date.strftime("%d %b %Y").upper()))
                wfile.write("\n")
                wfile.write("/\n\n")

                if any(currentcont=="WELSPECS"):
                    indices = np.logical_and(currentdate,welspec)
                    wfile.write("WELSPECS\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="COMPDATMD"):
                    indices = np.logical_and(currentdate,compdat)
                    wfile.write("COMPDATMD\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="COMPORD"):
                    indices = np.logical_and(currentdate,compord)
                    wfile.write("COMPORD\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WCONHIST"):
                    indices = np.logical_and(currentdate,prodhst)
                    wfile.write("WCONHIST\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WCONINJH"):
                    indices = np.logical_and(currentdate,injdhst)
                    wfile.write("WCONINJH\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WEFAC"):
                    indices = np.logical_and(currentdate,wefffac)
                    wfile.write("WEFAC\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WELOPEN"):
                    indices = np.logical_and(currentdate,welopen)
                    wfile.write("WELOPEN\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

class VTKit(DirBase):

    def __init__(self):

        pass

    def read(self,):

        pass

    def write(self,):

        pass

# Supporting String Classes

class Alphabet():

    aze_cyril_lower = [
        "а","б","ҹ","ч","д","е","я","ф","ҝ","ғ","һ","х","ы","и","ж","к",
        "г","л","м","н","о","ю","п","р","с","ш","т","у","ц","в","й","з"]

    aze_latin_lower = [
        "a","b","c","ç","d","e","ə","f","g","ğ","h","x","ı","i","j","k",
        "q","l","m","n","o","ö","p","r","s","ş","t","u","ü","v","y","z"]

    aze_cyril_upper = [
        "А","Б","Ҹ","Ч","Д","Е","Я","Ф","Ҝ","Ғ","Һ","Х","Ы","И","Ж","К",
        "Г","Л","М","Н","О","Ю","П","Р","С","Ш","Т","У","Ц","В","Й","З"]

    aze_latin_upper = [
        "A","B","C","Ç","D","E","Ə","F","G","Ğ","H","X","I","İ","J","K",
        "Q","L","M","N","O","Ö","P","R","S","Ş","T","U","Ü","V","Y","Z"]

    def __init__(self,string):

        self.string = string

    def convert(self,language="aze",from_="cyril",to="latin"):

        from_lower = getattr(self,f"{language}_{from_}_lower")
        from_upper = getattr(self,f"{language}_{from_}_upper")

        to_lower = getattr(self,f"{language}_{to}_lower")
        to_upper = getattr(self,f"{language}_{to}_upper")

        for from_letter,to_letter in zip(from_lower,to_lower):
            self.string.replace(from_letter,to_letter)

        for from_letter,to_letter in zip(from_upper,to_upper):
            self.string.replace(from_letter,to_letter)

if __name__ == "__main__":

    import unittest

    from tests import datatest

    unittest.main(datatest)