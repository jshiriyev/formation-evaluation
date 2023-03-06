import contextlib
import copy
import datetime

from difflib import SequenceMatcher

import logging

import os
import re

import numpy
import openpyxl

from .folder._browser import Browser

from .datum.frame._builtin import Frame

class XlBook(Browser):

    def __init__(self,**kwargs):

        super().__init__(**kwargs)

        self.sheets = {}

    def split(self,sheetname,transpose=False,intFlag=False):
        """It should split to frames based on all None lists."""

        frames = []
        
        subrows,prevRowNoneFlag = [],True

        for row in self.sheets[sheetname]:

            unwanted_element_count = countnone(row)

            if intFlag:
                unwanted_element_count += countint(row)

            if unwanted_element_count==len(row):
                if not prevRowNoneFlag:
                    frames.append(Frame(subrows))
                subrows,prevRowNoneFlag = [],True

            else:
                subrows.append(row)
                prevRowNoneFlag = False

        if not transpose:
            return frames

        for index,frame in enumerate(frames):

            frame = frame.transpose()
            frame = frame.drop()

            frames[index] = frame

        return frames

    def merge(self,cols=None,idframes=None,infosearch=False):
        """It merges all the frames as a single frame under the Excel class."""

        framemerged = frame()

        if cols is None:
            pass
        elif type(cols) is int:
            cols = (cols,)
        elif type(cols) is str:
            cols = (cols,)
        else:
            cols = tuple(cols)

        if idframes is None:
            idframes = range(len(self.frames))
        elif type(idframes) is int:
            idframes = (idframes,)
        elif type(idframes) is str:
            idframes = (idframes,)
        else:
            idframes = tuple(idframes)

        for idframe in idframes:

            frame = self.frames[idframe]

            if cols is None:
                datacolumns = [datacolumn for datacolumn in dataframe.running]
            elif not infosearch:
                datacolumns = []
                for index in tuple(cols):
                    if type(index) is int:
                        datacolumns.append(dataframe.running[index])
                    elif type(index) is str:
                        datacolumns.append(dataframe[index])
                    else:
                        raise TypeError(f"cols type should be either int or str, not {type(index)}")
            else:
                datacolumns = []
                for index in tuple(cols):
                    if type(index) is str:
                        scores = [SequenceMatcher(None,info,index).ratio() for info in frame.infos]
                        datacolumns.append(dataframe.running[scores.index(max(scores))])
                    else:
                        raise TypeError(f"cols type should be str, not {type(index)}")

            framemerged._append(*datacolumns)

        return framemerged

    def write(self,filepath,title):

        wb = openpyxl.Workbook()

        sheet = wb.active

        if title is not None:
            sheet.title = title

        for line in running:
            sheet.append(line)

        wb.save(filepath)

def loadxl(*args,sheetnames=None,**kwargs):
    """
    Returns an instance of textio.XlBook. If a filepath is specified, the instance
    represents the file.
    
    Arguments:
        filepath {str} -- path to the given excel file

    Keyword Arguments:
        homedir {str} -- path to the home (output) directory
        filedir {str} -- path to the file (input) directory
    
    Returns:
        textio.XlBook -- an instance of textio.XlBook filled with excel file text.

    """

    if len(args)==1:
        filepath = args[0]
    elif len(args)>1:
        raise "The function does not take more than one positional argument."

    # It creates an empty textio.XlBook instance.
    nullbook = XlBook(filepath=filepath,**kwargs)

    # It reads excel file and returns textio.XlBook instance.
    fullbook = XlWorm(nullbook,sheetnames=sheetnames).xlbook

    return fullbook

class XlWorm():

    def __init__(self,xlbook,sheetnames=None,**kwargs):

        self.xlbook = xlbook

        search = pop(kwargs,"search",False)

        min_row = pop(kwargs,"min_row",1)
        min_col = pop(kwargs,"min_col",1)

        max_row = pop(kwargs,"max_row",None)
        max_col = pop(kwargs,"max_col",None)

        with XlWorm.xlopen(self.xlbook.filepath) as self.book:

            if sheetnames is None:
                sheetnames = self.book.sheetnames
            else:
                sheetnames = self.get_sheetname(name=sheetnames,search=search)

            for sheetname in sheetnames:

                print("Loading {} {}".format(self.xlbook.filepath,sheetname))

                rows = self.load(sheetname,min_row,min_col,max_row,max_col)

                self.xlbook.sheets[sheetname] = rows

    def get_sheetname(self,name=None,search=False):

        if isinstance(name,int):
            sheetnames = [self.book.sheetnames[name]]
        elif isinstance(name,str) and search:
            mathcscore = [SequenceMatcher(None,sheetname,name).ratio() for sheetname in self.book.sheetnames]
            sheetnames = [self.book.sheetnames[mathcscore.index(max(mathcscore))]]
        elif isinstance(name,str):
            if name not in self.book.sheetnames:
                raise ValueError(f"'{name}' could not be found in the xlbook, try search=True.")
            sheetnames = [name]
        elif hasattr(name,"len"):
            sheetnames = [self.get_sheetname(name=sheetname,search=search) for sheetname in name]
        else:
            raise TypeError(f"Expected sheetnames is either none, int or str, but the input type is {type(sheetnames)}.")

        return sheetnames

    def load(self,sheetname,min_row=1,min_col=1,max_row=None,max_col=None):
        """It reads provided excel worksheet and returns it as a frame."""

        rows = self.book[sheetname].iter_rows(
            min_row=min_row,max_row=max_row,
            min_col=min_col,max_col=max_col,
            values_only=True)

        return [list(row) for row in rows]

    @contextlib.contextmanager
    def xlopen(filepath):
        xlbook = openpyxl.load_workbook(filepath,read_only=True,data_only=True,keep_links=False)
        try:
            yield xlbook
        finally:
            xlbook._archive.close()

def pop(kwargs,key,default=None):

    try:
        return kwargs.pop(key)
    except KeyError:
        return default

def countnone(_list):

    return _list.count(None)

def countint(_list):

    counter = 0

    for value in _list:

        if isinstance(value,int):
            counter += 1

    return counter